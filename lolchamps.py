from xmlrpc.client import MAXINT, MININT
import discord
from discord_components import DiscordComponents, Button
import random
import string
import cassiopeia as cass
import psycopg2
import requests
import re

from openpyxl import load_workbook

import os
from dotenv import load_dotenv

load_dotenv()

TOKEN = os.getenv('TOKEN')
DATABASE_URL = os.getenv('DATABASE_URL')
con = psycopg2.connect(DATABASE_URL)
cur = con.cursor()

clp = requests.get('https://ddragon.leagueoflegends.com/api/versions.json').json()[0] #CurrentLeaguePatch
champs = requests.get(f'https://ddragon.leagueoflegends.com/cdn/{clp}/data/en_US/champion.json').json()["data"] #CurrentLeaguePatch

urls = {}
champnames = {}

for champ in champs.values():
    name=champ["name"]
    key=champ["id"]
    champnames[champ["key"]]=name
    splash=f'https://ddragon.leagueoflegends.com/cdn/img/champion/splash/{key}_0.jpg'
    urls[name]=splash

urls["summoners"]="https://cdn.discordapp.com/attachments/518907821081755672/987335003371364452/unknown.png"


rawchampselect = os.popen("curl https://raw.communitydragon.org/latest/plugins/rcp-fe-lol-champion-statistics/global/default/rcp-fe-lol-champion-statistics.js").read()
rawplayrates = (rawchampselect.split("a.exports=")[1].split("},function")[0])
rawlaners = rawplayrates.replace("TOP","|TOP").replace("JUNGLE","|JGL").replace("MIDDLE","|MID").replace("BOTTOM","|ADC").replace("SUPPORT","|SUP").replace("{","").replace("}","").split("|")[1:]
laners = []
currentlane = 0
for lane in rawlaners:
    
    betterlane = "," + lane[4:]
    # print(betterlane)
    currentlane = [lane[:3]] + re.findall(',(.*?):',betterlane)
    laners.append(currentlane)

laners[4],laners[3]=laners[3],laners[4]

lanes = {'all': 8, 'top': 9, 'jg': 10, 'jung': 10, 'jng': 10,'jungle': 10, 'jungler': 10, 'mid': 11, 'adc': 12, 'bot': 12, 'supp': 13, 'sup': 13}

def generate_embed(champ,lane):
    embedVar = discord.Embed(color=0x00ff00)
    embedVar.add_field(name="Champion", value=champ)
    embedVar.add_field(name="Linea", value=lane)
    embedVar.set_image(url=urls[champ])
    return embedVar

def cords_to_cellname(row,col):
    alph = string.ascii_uppercase
    return(alph[col-1]+str(row))

def random_champ(wks, lane):
    laneid = lanes[lane]
    limit = 0
    for cell in wks[string.ascii_uppercase[laneid-1]]:
        if cell.value != None:
            limit+=1

    if limit == 2: return "Azir"
    return (wks.cell(random.randint(3, limit), laneid).value)

def remove_champ(wks, champ):
    champs = []
    for ch in wks[string.ascii_uppercase[7]]:
        champs.append(ch.value)

    if not champ in champs:
        return

    cols = ['H', 'I', 'J', 'K', 'L', 'M']
    cell_list = []

    for col in cols:
        for cell in wks[col]:
            if cell.value == champ:
                cell_list.append(cell)

    while len(cell_list) > 0:
        currentrow = cell_list[-1].row
        currentcol = cell_list[-1].column
        if currentcol < 8:
            cell_list.pop()
            continue
        limit = 0
        for cell in wks[string.ascii_uppercase[currentcol-1]]:
            if cell.value != None:
                limit+=1
        

        listofchamps = []
        for i in range(currentrow+1,limit+1):
            listofchamps.append(wks.cell(row=i,column=currentcol).value)

        j = currentrow
        for i in listofchamps:
            wks.cell(row=j,column=currentcol,value=i)
            j+=1
        wks.cell(row=limit,column=currentcol,value='')
        cell_list.pop() 

async def printlaner(wks,lane,champmsg,interaction,components):
    champ = random_champ(wks,lane)
    embedVar = generate_embed(champ, lane)
    await champmsg.edit('',embed=embedVar, **components)
    await interaction.send(content=f"<a:kirby:759485375718359081>Re-Roll {lane}<a:kirby:759485375718359081>",ephemeral=False, delete_after=2)
    
class MyClient(discord.Client):

    async def on_ready(self):
        print('Logged on as', self.user)

    async def on_message(self, message):
        if (self.user in message.mentions) or (f"<@&{self.user.id}>" in message.content):
            pid = random.randint(MININT,MAXINT)
            components = {"components" : [[Button(label="Win", style="3", emoji = self.get_emoji(id=987155911766335520), custom_id=f"win{pid}"), 
            Button(label="Re-Roll", style="1", emoji = "🔁", custom_id=f"roll{pid}"), Button(label="Delete", style="4", emoji = self.get_emoji(id=987331408093642822), custom_id=f"del{pid}"),
            Button(label="Lanes", style="2", emoji =self.get_emoji(id=987173438907088966) , custom_id=f"lanes{pid}")]]}
            lanebuttons = [Button(emoji = self.get_emoji(id=987155912890417154),custom_id=f"top{pid}"),Button(emoji = self.get_emoji(id=987155914362589184),custom_id=f"jg{pid}"),
            Button(emoji = self.get_emoji(id=987155915541205002),custom_id=f"mid{pid}"),Button(emoji = self.get_emoji(id=987172387277656104),custom_id=f"adc{pid}"),
            Button(emoji = self.get_emoji(id=987155917961318440),custom_id=f"supp{pid}")]
            lane = message.clean_content.replace(f"@{self.user.name} ", '').lower()
            knownuser = True
            buttons = {f"roll{pid}":0, f"win{pid}":1, f"del{pid}":2, f"lanes{pid}":3,f"top{pid}":4,f"jg{pid}":5,f"mid{pid}":6,f"adc{pid}":7,f"supp{pid}":8,f"yescreate{pid}":9,f"nocreate{pid}":10}
            lanembed = discord.Embed(color=0x00ff00).set_image(url=urls["summoners"])
            wb = load_workbook('lolchamps.xlsx')
            wks = wb["Clean"]
            if not str(message.author.id) in wb.sheetnames:
                knownuser = False
            else: 
                wks = wb[str(message.author.id)]
            if lane in lanes:
                champ = random_champ(wks,lane)
                embedVar = generate_embed(champ,lane)
                champmsg = await message.reply(embed=embedVar, **components)

            else:
                champmsg = await message.reply('Que linea quieres bro', embed=lanembed, components=[lanebuttons])

            while True:
                try:
                    interaction = await client.wait_for("button_click")
                    customid=interaction.component.custom_id

                    match buttons[customid]:
                        case 0: #Button for re-rolling
                            pass
                        case 1: #Button for selecting a win
                            if interaction.user != message.author:
                                await interaction.send("No es tu campeón amigo")
                            elif knownuser:
                                wks = wb[str(message.author.id)]
                                champ = interaction.message.content.rsplit(' ', 1)[0]
                                remove_champ(wks,champ)
                                wb.save('lolchamps.xlsx')
                                await champmsg.delete()
                                await message.delete()
                                await interaction.send("Quitado de la lista")
                            else:
                                confirmmsg = await interaction.send("No tienes ficha creada. ¿Quieres crear una?", components[[Button(label="Yes", style="3", emoji = self.get_emoji(id=987155911766335520), custom_id=f"yescreate{pid}"),
                                Button(label="No", style="4", emoji = self.get_emoji(id=987155911766335520), custom_id=f"nocreate{pid}")]])
                            continue
                        case 2: #Button for deleting the messages
                            await champmsg.delete()
                            await message.delete()
                            break
                        case 3: #Button for selecting lanes
                            await champmsg.edit('Que linea quieres bro', embed=lanembed, components=[lanebuttons])
                            await interaction.send(content="<a:kirby:759485375718359081>Lineas<a:kirby:759485375718359081>",ephemeral=False, delete_after=2)
                            continue
                        case 4: #Button for selecting top
                            lane = "top"
                        case 5: #Button for selecting jungle
                            lane = "jungle"
                        case 6: #Button for selecting mid
                            lane = "mid"
                        case 7: #Button for selecting ADCarry
                            lane = "adc"
                        case 8: #Button for selecting support
                            lane = "supp"
                        case 9: #Button for creating a table for the user
                            # cur.execute(f"""CREATE TABLE {interaction.user}_table AS SELECT *
                            #             FROM clean_table""")
                            # con.commit()
                            await confirmmsg.delete()
                        case 10: ##Button for not creating a table for the user
                            await confirmmsg.delete()

                    await printlaner(wks,lane,champmsg,interaction,components)
                except:
                    await message.channel.send("Exception", delete_after=10)
                    await champmsg.delete()
                    await message.delete()
                    break


cur.execute("select exists(select * from information_schema.tables where table_name=%s)", ('clean_table',))
if not (cur.fetchone()[0]):
    cur.execute("""CREATE TABLE clean_table (
    CHAMP VARCHAR(255),
    ID INTEGER, 
    TOP BOOLEAN, 
    JGL BOOLEAN, 
    MID BOOLEAN, 
    ADC BOOLEAN, 
    SUP BOOLEAN);""")

    for id in list(champnames.keys()):
        l = lanes(id)
        cur.execute("INSERT INTO clean_table(CHAMP,ID,TOP,JGL,MID,ADC,SUP) VALUES (%s, %s,%s,%s,%s,%s,%s)", (champnames[id],id,l[0],l[1],l[2],l[3],l[4],))
    con.commit()


intents = discord.Intents(messages=True, guilds=True)
client = MyClient(intents=intents)
DiscordComponents(client)
client.run(str(TOKEN))